# import openpyxl, smtplib
# from pathlib import Path
# """
# تحميل ملف إكسل يحتوي على بيانات الأعضاء، والتحقق من الأعضاء
# الذين لم يدفعوا مستحقاتهم الشهرية، ثم إرسال بريد إلكتروني تذكيري لهم
# """
# # connect with Excelfile
# file = openpyxl. load_workbook(Path.home() / Path('Desktop', 'members.xlsx'))
# sheets = file.sheetnames
# sheet = file['Sheet1']
# lastCol = sheet.max_column
# latestMonth = sheet.cell(rom=1, column=lastCol).value
# print(latestMonth)

# #قاموس للي لم يدفعو المستقات
# unpaidMembers = {}

# for r in range(2, sheet.max_row + 1):
#     payment = sheet.cell(rom=r, column=lastCol).value

#     if payment != 'paid':
#         name = sheet.cell(rom=r, column=1).value
#         email = sheet.cell(rom=r, column=2).value
#         unpaidMembers[name] = email

# # ارسال ايميل
# smtp0bj = smtplib.SMTP('smtp.gmail.com', 587)
# smtp0bj.starttls()
# sender_email = "academyhsoub1@gmail.com"
# password = input(str("Please enter your password: "))
# smtp0bj.login(sender_email, password)

# for name, email in unpaidMembers.items():
#     body = """Subject: %s dues unpaid. \nDear %s, \nRecords show that you have not paid dues
#     for %s. Please make this payment as soon as possible. Thank you!'""" %(latestMonth, name, latestMonth)
#     print('Sending email to %s ... ' % email)
#     sendmailStatus = smtp0bj.sendmail(sender_email, email, body)

#     if sendmailStatus != {}:
#         print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

# smtp0bj.quit()
import openpyxl
from openpyxl import Workbook
from pathlib import Path

# إنشاء ملف إكسل جديد
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# إضافة الترويسة
headers = ["Name", "Email", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
ws.append(headers)

# إضافة بعض البيانات الافتراضية
members = [
    ["Alice Johnson", "alice@example.com", "paid", "paid", "unpaid", "paid", "paid", "unpaid", "paid", "paid", "paid", "unpaid", "paid", "paid"],
    ["Bob Smith", "bob@example.com", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "unpaid", "unpaid", "paid", "paid", "unpaid"],
    ["Charlie Brown", "charlie@example.com", "unpaid", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "paid"],
    ["David Wilson", "david@example.com", "paid", "paid", "paid", "unpaid", "paid", "unpaid", "paid", "paid", "paid", "paid", "paid", "unpaid"],
    ["Eva Green", "eva@example.com", "paid", "unpaid", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "paid", "unpaid", "paid"]
]

for member in members:
    ws.append(member)

# حفظ الملف في سطح المكتب
file_path = Path.home() / Path('Desktop', 'members.xlsx')
wb.save(file_path)
print(f"File saved to {file_path}")
