import openpyxl, smtplib, datetime
""" 
من الاكسل نستخرج المعلومات ونرسل رساله بالايميل للي ما دفع الايجار
الجدول الي بالاكسل:

		   |                      |Jan-23	|Feb-23	|Mar-23	|Apr-23	|May-23
ali        |alshihy.567a@gmail.com|	pay	    |pay	|		|       |pay
momammed   |alshihy.567a@gmail.com|	pay	    |pay	|pay    |pay	|pay
maram      |alshihy.567a@gmail.com|	pay		|       |pay 	|pay	|pay
alawi      |alshihy.567a@gmail.com|	pay	    |pay	|   	|pay	|pay

"""
file = openpyxl.load_workbook("people.xlsx")

sheet = file.active


#طول الاعمدة والصفوف
max_rowX = sheet.max_row
max_columnX = sheet.max_column

dont_pay = {}
#مر على صفوف
for i in range(2, max_rowX + 1):
    #يبدا اللست بايميل المرسل له
    #بعدين الوقت الي ما دفع فيه
    #عند استخراج التاريخ من اكسل يكون من نوع ديتتايم
    dont_pay_list = []

    #مر على الاعمدة
    for j in range(3, max_columnX + 1):
        cell = sheet.cell(row=i, column=j).value
        print(cell)
        #تاكد انه لم يدفع
        if cell != "pay":
            #تاكد ان ايميل الشخص موجود تضيفه اول اندكس
            emai_per = sheet.cell(row=i, column=2).value
            if emai_per not in dont_pay_list:
                dont_pay_list.append(emai_per)
            dont_pay_list.append(sheet.cell(row=1, column=j).value)


    if dont_pay_list != []:
        dont_pay[sheet.cell(row=i, column=1).value] = dont_pay_list

print(dont_pay)


email_from = "alshihy.567a@gmail.com"
password = "jqbh kcvj sdhe jdai"

server = smtplib.SMTP("smtp.gmail.com", 587)#تفعيل سرفر
server.starttls()#تشفير
#تسجيل دخول
server.login(email_from, password)

for peaple in dont_pay:
    print(str(dont_pay[peaple][0]) , "Subject:dont pay\ndear " + peaple + " you net to pay your" + str(dont_pay[peaple][1:]))

    user = str(dont_pay[peaple][0])
    # strftime تخلي نوع الدي تايم لسترنج حسب التسيق المختار
    massage = "ٍ Subject:Rent\n " + str(peaple) + " you are not pay your rent at:" + "".join("\n- "+str(i.strftime("%d/ %b/ %y")) for i in dont_pay[peaple][1:])
    #يطبع النتائج
    print("user: " + user)
    print("massage: " + massage)
    print()
    #يرسل النتائج
    # server.sendmail(email_from, user, massage)
server.quit()
